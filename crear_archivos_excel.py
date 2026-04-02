#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para crear archivos Excel de mantenimiento automáticamente
Genera 25 archivos Excel con información de cada mantenimiento
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import random
from datetime import datetime, timedelta

# Datos de los mantenimientos
mantenimientos = [
    {
        "id": 1, "equipo": "AGITADOR CIRCULAR VIBRATORIO PARA TUBOS", "codigo": "EQ-001",
        "fecha": "2024-01-15", "tecnico": "Gabriel Leon", "tipo": "Preventivo",
        "descripcion": "Mantenimiento preventivo mensual. Limpieza de componentes, calibración de sensores y verificación de funcionamiento.",
        "costo": 150.00, "estado": "Completado"
    },
    {
        "id": 2, "equipo": "AGITADOR TIPO VORTEX PARA MICROTUBOS", "codigo": "EQ-002",
        "fecha": "2024-01-20", "tecnico": "Santiago Nuñez", "tipo": "Preventivo",
        "descripcion": "Reemplazo de filtros y lubricación de partes móviles. Verificación de presión y temperatura.",
        "costo": 200.00, "estado": "Completado"
    },
    {
        "id": 3, "equipo": "AGLUTINOSCOPIO PARA ANÁLISIS DE SANGRE", "codigo": "EQ-003",
        "fecha": "2024-02-05", "tecnico": "Kerly Verdezoto", "tipo": "Correctivo",
        "descripcion": "Mantenimiento correctivo. Reparación de sistema de agitación y calibración de sensores.",
        "costo": 350.00, "estado": "Completado"
    },
    {
        "id": 4, "equipo": "ANALIZADOR DE BIOQUÍMICA AUTOMATIZADO", "codigo": "EQ-004",
        "fecha": "2024-02-10", "tecnico": "Amy Aguagiña", "tipo": "Preventivo",
        "descripcion": "Actualización de software y calibración de analizadores. Verificación de precisión.",
        "costo": 180.00, "estado": "Completado"
    },
    {
        "id": 5, "equipo": "ANALIZADOR DE ELECTROLITOS DIAMOND", "codigo": "EQ-005",
        "fecha": "2024-02-25", "tecnico": "Gabriel Leon", "tipo": "Preventivo",
        "descripcion": "Mantenimiento preventivo trimestral. Limpieza profunda y reemplazo de consumibles.",
        "costo": 280.00, "estado": "Completado"
    },
    {
        "id": 6, "equipo": "ANALIZADOR HEMATOLÓGICO MICROS ES 60", "codigo": "EQ-006",
        "fecha": "2024-03-08", "tecnico": "Santiago Nuñez", "tipo": "Correctivo",
        "descripcion": "Reparación de sistema de refrigeración y calibración de termostatos.",
        "costo": 420.00, "estado": "Completado"
    },
    {
        "id": 7, "equipo": "ARCO DENTAL ASCENSOR CIGOMÁTICO", "codigo": "EQ-007",
        "fecha": "2024-03-15", "tecnico": "Kerly Verdezoto", "tipo": "Preventivo",
        "descripcion": "Mantenimiento preventivo. Limpieza de lentes y calibración de sistema óptico.",
        "costo": 160.00, "estado": "Completado"
    },
    {
        "id": 8, "equipo": "ARMÓNICO ULTRASÓNICO PARA CIRUGÍA", "codigo": "EQ-008",
        "fecha": "2024-03-22", "tecnico": "Amy Aguagiña", "tipo": "Preventivo",
        "descripcion": "Reemplazo de piezas desgastadas y lubricación de mecanismos.",
        "costo": 190.00, "estado": "Completado"
    },
    {
        "id": 9, "equipo": "ASPIRADOR DE FLUIDOS QUIRÚRGICOS", "codigo": "EQ-009",
        "fecha": "2024-04-05", "tecnico": "Gabriel Leon", "tipo": "Correctivo",
        "descripcion": "Mantenimiento correctivo. Reparación de sistema de succión y filtros.",
        "costo": 380.00, "estado": "Completado"
    },
    {
        "id": 10, "equipo": "AUDIÓMETRO CLÍNICO DE DIAGNÓSTICO", "codigo": "EQ-010",
        "fecha": "2024-04-12", "tecnico": "Santiago Nuñez", "tipo": "Preventivo",
        "descripcion": "Calibración de sistema auditivo y verificación de frecuencias.",
        "costo": 220.00, "estado": "Completado"
    },
    {
        "id": 11, "equipo": "AUTOCLAVE DE MESA AUTOMÁTICO", "codigo": "EQ-011",
        "fecha": "2024-04-25", "tecnico": "Kerly Verdezoto", "tipo": "Preventivo",
        "descripcion": "Mantenimiento preventivo semestral. Limpieza profunda y reemplazo de juntas.",
        "costo": 320.00, "estado": "Completado"
    },
    {
        "id": 12, "equipo": "AUTOCLAVE DE SOBREMESA 20-30 LITROS", "codigo": "EQ-012",
        "fecha": "2024-05-03", "tecnico": "Amy Aguagiña", "tipo": "Preventivo",
        "descripcion": "Actualización de software y calibración de sensores de presión.",
        "costo": 170.00, "estado": "Completado"
    },
    {
        "id": 13, "equipo": "AUTORREFRACTOMETRO AUTOMÁTICO", "codigo": "EQ-013",
        "fecha": "2024-05-10", "tecnico": "Gabriel Leon", "tipo": "Correctivo",
        "descripcion": "Mantenimiento correctivo. Reparación de sistema de enfoque automático.",
        "costo": 450.00, "estado": "Completado"
    },
    {
        "id": 14, "equipo": "BALANZA DE PISO CON TALLÍMETRO", "codigo": "EQ-014",
        "fecha": "2024-05-18", "tecnico": "Santiago Nuñez", "tipo": "Preventivo",
        "descripcion": "Calibración de balanza y verificación de precisión de medición.",
        "costo": 140.00, "estado": "Completado"
    },
    {
        "id": 15, "equipo": "BAÑO DE PARAFINA PORTÁTIL", "codigo": "EQ-015",
        "fecha": "2024-05-25", "tecnico": "Kerly Verdezoto", "tipo": "Preventivo",
        "descripcion": "Mantenimiento preventivo. Limpieza de sistema de calefacción.",
        "costo": 180.00, "estado": "Completado"
    },
    {
        "id": 16, "equipo": "BAÑO MARÍA DIGITAL TEMPERATURA", "codigo": "EQ-016",
        "fecha": "2024-06-05", "tecnico": "Amy Aguagiña", "tipo": "Preventivo",
        "descripcion": "Reemplazo de termostatos y calibración de temperatura.",
        "costo": 210.00, "estado": "Completado"
    },
    {
        "id": 17, "equipo": "BICICLETA ESTÁTICA DE EJERCICIO", "codigo": "EQ-017",
        "fecha": "2024-06-12", "tecnico": "Gabriel Leon", "tipo": "Correctivo",
        "descripcion": "Mantenimiento correctivo. Reparación de sistema de pedaleo.",
        "costo": 290.00, "estado": "Completado"
    },
    {
        "id": 18, "equipo": "BLOQUE SECO PARA TUBOS DE ENSAYO", "codigo": "EQ-018",
        "fecha": "2024-06-20", "tecnico": "Santiago Nuñez", "tipo": "Preventivo",
        "descripcion": "Limpieza de sistema de calefacción y verificación de temperatura.",
        "costo": 160.00, "estado": "Completado"
    },
    {
        "id": 19, "equipo": "BOMBA DE INFUSIÓN PERISTÁLTICA", "codigo": "EQ-019",
        "fecha": "2024-06-28", "tecnico": "Kerly Verdezoto", "tipo": "Preventivo",
        "descripcion": "Mantenimiento preventivo. Calibración de sistema de infusión.",
        "costo": 240.00, "estado": "Completado"
    },
    {
        "id": 20, "equipo": "CABINA DE SEGURIDAD BIOLÓGICA CLASE II", "codigo": "EQ-020",
        "fecha": "2024-07-05", "tecnico": "Amy Aguagiña", "tipo": "Preventivo",
        "descripcion": "Mantenimiento preventivo trimestral. Limpieza de filtros HEPA y verificación de flujo de aire.",
        "costo": 380.00, "estado": "Completado"
    },
    {
        "id": 21, "equipo": "AGITADOR CIRCULAR VIBRATORIO PARA TUBOS", "codigo": "EQ-001",
        "fecha": "2024-07-12", "tecnico": "Gabriel Leon", "tipo": "Preventivo",
        "descripcion": "Mantenimiento preventivo mensual. Verificación de funcionamiento y limpieza.",
        "costo": 150.00, "estado": "Completado"
    },
    {
        "id": 22, "equipo": "AGLUTINOSCOPIO PARA ANÁLISIS DE SANGRE", "codigo": "EQ-003",
        "fecha": "2024-07-20", "tecnico": "Santiago Nuñez", "tipo": "Correctivo",
        "descripcion": "Reparación de sistema de calibración y actualización de software.",
        "costo": 520.00, "estado": "Completado"
    },
    {
        "id": 23, "equipo": "ANALIZADOR DE ELECTROLITOS DIAMOND", "codigo": "EQ-005",
        "fecha": "2024-07-28", "tecnico": "Kerly Verdezoto", "tipo": "Preventivo",
        "descripcion": "Mantenimiento preventivo. Limpieza de componentes y calibración.",
        "costo": 200.00, "estado": "Completado"
    },
    {
        "id": 24, "equipo": "ARCO DENTAL ASCENSOR CIGOMÁTICO", "codigo": "EQ-007",
        "fecha": "2024-08-05", "tecnico": "Amy Aguagiña", "tipo": "Preventivo",
        "descripcion": "Reemplazo de lámpara y calibración de sistema óptico.",
        "costo": 180.00, "estado": "Completado"
    },
    {
        "id": 25, "equipo": "AUDIÓMETRO CLÍNICO DE DIAGNÓSTICO", "codigo": "EQ-010",
        "fecha": "2024-08-12", "tecnico": "Gabriel Leon", "tipo": "Correctivo",
        "descripcion": "Mantenimiento correctivo. Reparación de sistema de audio y calibración.",
        "costo": 410.00, "estado": "Completado"
    }
]

# Actualizar los datos de los mantenimientos para incluir los nuevos campos
for m in mantenimientos:
    # Simular fecha de falla y reparación
    fecha_falla = datetime.strptime(m['fecha'], '%Y-%m-%d') - timedelta(hours=random.randint(2, 48))
    fecha_reparacion = datetime.strptime(m['fecha'], '%Y-%m-%d')
    duracion_reparacion = (fecha_reparacion - fecha_falla).total_seconds() / 3600
    # Simular tiempo de funcionamiento desde la última falla (entre 100 y 1000 horas)
    tiempo_funcionamiento = random.randint(100, 1000)
    m['fecha_falla'] = fecha_falla.strftime('%Y-%m-%d %H:%M')
    m['fecha_reparacion'] = fecha_reparacion.strftime('%Y-%m-%d %H:%M')
    m['duracion_reparacion'] = round(duracion_reparacion, 2)
    m['tiempo_funcionamiento'] = tiempo_funcionamiento

def crear_archivo_excel(mantenimiento):
    """Crear un archivo Excel para un mantenimiento específico"""
    
    # Crear workbook
    wb = Workbook()
    
    # Hoja 1: Información General
    ws1 = wb.active
    ws1.title = "Información General"
    
    # Estilo para títulos
    titulo_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12)
    normal_font = Font(size=11)
    
    # Título principal
    ws1['A1'] = "REPORTE DE MANTENIMIENTO"
    ws1['A1'].font = titulo_font
    
    # Información del mantenimiento
    ws1['A3'] = f"Equipo: {mantenimiento['equipo']}"
    ws1['A4'] = f"Código: {mantenimiento['codigo']}"
    ws1['A5'] = f"Fecha de Mantenimiento: {mantenimiento['fecha']}"
    ws1['A6'] = f"Técnico Responsable: {mantenimiento['tecnico']}"
    ws1['A7'] = f"Tipo de Mantenimiento: {mantenimiento['tipo']}"
    ws1['A8'] = f"Estado: {mantenimiento['estado']}"
    ws1['A9'] = f"Costo: ${mantenimiento['costo']:.2f}"
    
    # Agregar los nuevos campos en la hoja 1
    ws1['A13'] = 'Fecha de Falla:'
    ws1['B13'] = mantenimiento['fecha_falla']
    ws1['A14'] = 'Fecha de Reparación:'
    ws1['B14'] = mantenimiento['fecha_reparacion']
    ws1['A15'] = 'Duración de Reparación (h):'
    ws1['B15'] = mantenimiento['duracion_reparacion']
    ws1['A16'] = 'Tiempo de Funcionamiento (h):'
    ws1['B16'] = mantenimiento['tiempo_funcionamiento']
    
    # Aplicar estilos
    for i in range(3, 10):
        ws1[f'A{i}'].font = normal_font
    
    # Descripción
    ws1['A11'] = "DESCRIPCIÓN DEL MANTENIMIENTO"
    ws1['A11'].font = header_font
    ws1['A12'] = mantenimiento['descripcion']
    ws1['A12'].font = normal_font
    
    # Hoja 2: Actividades Realizadas
    ws2 = wb.create_sheet("Actividades")
    
    ws2['A1'] = "ACTIVIDADES REALIZADAS"
    ws2['A1'].font = titulo_font
    
    # Encabezados
    headers = ["No.", "Actividad", "Estado", "Observaciones"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = header_font
    
    # Actividades de ejemplo
    actividades = [
        ["1", "Limpieza de componentes", "Completado", "Se limpiaron todos los componentes principales"],
        ["2", "Calibración de sensores", "Completado", "Sensores calibrados según especificaciones"],
        ["3", "Verificación de funcionamiento", "Completado", "Equipo funciona correctamente"],
        ["4", "Reemplazo de filtros", "Completado", "Filtros reemplazados por nuevos"],
        ["5", "Lubricación de partes móviles", "Completado", "Lubricación aplicada según manual"]
    ]
    
    for row, actividad in enumerate(actividades, 4):
        for col, valor in enumerate(actividad, 1):
            cell = ws2.cell(row=row, column=col, value=valor)
            cell.font = normal_font
    
    # Hoja 3: Materiales Utilizados
    ws3 = wb.create_sheet("Materiales")
    
    ws3['A1'] = "MATERIALES UTILIZADOS"
    ws3['A1'].font = titulo_font
    
    # Encabezados
    headers = ["No.", "Material", "Cantidad", "Costo Unitario", "Costo Total"]
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.font = header_font
    
    # Materiales de ejemplo
    materiales = [
        ["1", "Filtros de aire", "2", "$25.00", "$50.00"],
        ["2", "Aceite lubricante", "1", "$15.00", "$15.00"],
        ["3", "Paños de limpieza", "5", "$2.00", "$10.00"],
        ["4", "Consumibles varios", "1", "$75.00", "$75.00"],
        ["", "TOTAL", "", "", "$150.00"]
    ]
    
    for row, material in enumerate(materiales, 4):
        for col, valor in enumerate(material, 1):
            cell = ws3.cell(row=row, column=col, value=valor)
            cell.font = normal_font
    
    # Ajustar ancho de columnas
    for ws in [ws1, ws2, ws3]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb

def main():
    """Función principal para crear todos los archivos Excel"""
    
    # Crear directorio si no existe
    output_dir = "archivos_mantenimiento"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    print("🔧 Creando archivos Excel de mantenimiento...")
    
    for mantenimiento in mantenimientos:
        # Crear archivo Excel
        wb = crear_archivo_excel(mantenimiento)
        
        # Guardar archivo
        filename = f"mnt_{mantenimiento['id']:03d}.xlsx"
        filepath = os.path.join(output_dir, filename)
        wb.save(filepath)
        
        print(f"✅ Creado: {filename} - {mantenimiento['equipo'][:30]}...")
    
    print(f"\n🎉 ¡Completado! Se crearon {len(mantenimientos)} archivos Excel en la carpeta '{output_dir}'")
    print("📁 Ahora puedes subir estos archivos a Supabase Storage")

if __name__ == "__main__":
    main() 